import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 要保留的字段
KEEP_COLUMNS = [
    "结算日期", "司机", "司机名称", "计费流水号", "总金额",
    "配送费", "安装费", "日对账单号", "实际结算日期"
]

DRIVER_COLUMN = "司机名称"
AMOUNT_COLUMN = "总金额"

# 设置全局变量保存当前文件路径和数据
current_file_path = None
current_df = None

# 设置 Excel 列宽自适应 + 格式美化
def format_excel(filepath, driver_name, total_amount):
    wb = load_workbook(filepath)
    ws = wb.active

    # 删除第二行（多余的空行）
    ws.delete_rows(2)

    # 插入标题行
    ws.insert_rows(1)

    # 合并单元格
    ws.merge_cells("A1:C1")
    ws.merge_cells("D1:I1")

    # 设置样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True, size=20)

    # 设置司机名称
    ws["A1"] = driver_name
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = bold_font
    ws["A1"].fill = yellow_fill

    # 设置合计金额
    ws["D1"] = f"{total_amount:.2f}"
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["D1"].font = bold_font
    ws["D1"].fill = yellow_fill

    # 设置列宽自适应
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # 获取列号
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # 内容居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filepath)

# 加载文件并预处理
def load_file(file_path):
    global current_file_path, current_df
    try:
        df = pd.read_excel(file_path, dtype=str)
    except Exception as e:
        messagebox.showerror("读取失败", f"无法读取文件: {e}")
        return

    # 清理列名和数据空格
    df.columns = df.columns.str.strip().str.replace("\xa0", "", regex=True)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    # 检查字段是否存在
    missing = [col for col in KEEP_COLUMNS if col not in df.columns]
    if missing:
        messagebox.showerror("字段缺失", f"缺少以下字段: {', '.join(missing)}")
        return

    df = df[KEEP_COLUMNS]
    current_file_path = file_path
    current_df = df
    messagebox.showinfo("加载成功", "文件加载成功，请点击“核对金额”或“拆分”")

# 核对金额功能
def check_total_amount():
    if current_df is None:
        messagebox.showwarning("未加载文件", "请先加载一个文件")
        return

    try:
        total = current_df[AMOUNT_COLUMN].astype(float).sum()
    except:
        total = 0.0

    messagebox.showinfo("总金额核对", f"账单总金额为：{total:.2f}")

# 拆分功能
def split_file():
    if current_df is None or current_file_path is None:
        messagebox.showwarning("未加载文件", "请先加载一个文件")
        return

    grouped = current_df.groupby(DRIVER_COLUMN)
    output_dir = os.path.join(os.path.dirname(current_file_path), "拆分结果")
    os.makedirs(output_dir, exist_ok=True)

    for driver, group in grouped:
        if not driver:
            continue

        try:
            total_amount = group[AMOUNT_COLUMN].astype(float).sum()
        except:
            total_amount = 0.0

        filename = f"{driver}.xlsx"
        output_path = os.path.join(output_dir, filename)
        group.to_excel(output_path, index=False)

        format_excel(output_path, driver, total_amount)

    messagebox.showinfo("完成", f"拆分完成！文件保存在:\n{output_dir}")

# 拖拽支持（Windows）
def drop(event):
    file_path = event.data.strip('{}')  # 处理空格或路径中的大括号
    if file_path.endswith(('.xlsx', '.xls')):
        load_file(file_path)
    else:
        messagebox.showwarning("格式错误", "请拖入 Excel 文件")

# 文件选择按钮
def select_file():
    file_path = filedialog.askopenfilename(
        title="选择苏宁对账单 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx *.xls")]
    )
    if file_path:
        load_file(file_path)

# 主窗口
def main():
    root = tk.Tk()
    root.title("苏宁对账单拆分工具")
    root.geometry("500x300")

    label = tk.Label(root, text="拖拽文件到此窗口或点击选择文件", font=("Arial", 14))
    label.pack(pady=20)

    # 拖拽区域（Windows 下需安装 tkdnd）
    try:
        import tkinterdnd2 as tkdnd
        root = tkdnd.TkinterDnD.Tk()
        label = tk.Label(root, text="拖拽 Excel 文件到此区域", relief="groove", width=40, height=4)
        label.pack(pady=10)
        label.drop_target_register(tkdnd.DND_FILES)
        label.dnd_bind('<<Drop>>', drop)
    except ImportError:
        print("未安装 tkinterdnd2，拖拽功能不可用")

    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=20)

    btn_select = tk.Button(btn_frame, text="选择文件", command=select_file, width=15, height=2)
    btn_select.grid(row=0, column=0, padx=10)

    btn_check = tk.Button(btn_frame, text="核对金额", command=check_total_amount, width=15, height=2)
    btn_check.grid(row=0, column=1, padx=10)

    btn_split = tk.Button(btn_frame, text="拆分文件", command=split_file, width=15, height=2)
    btn_split.grid(row=0, column=2, padx=10)

    root.mainloop()

if __name__ == "__main__":
    main()
